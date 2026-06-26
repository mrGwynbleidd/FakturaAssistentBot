
#our main file
#read -> map
import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.importers.call_center_mapping import STANDARD_FIELDS, COLUMN_ALIASES

from app.importers.text_cleaner import clean_text

REQUITRD_FIELDS = [
    "question",
    "answer",
]

def normalize_key(value: Any) -> str:
    text = clean_text(value).lower()

    return (
        text
        .replace("ё", "е")
        .replace(" ", "_")
        .replace("-", "_")
        .strip()
    )

def build_alias_lookup() -> dict[str, str]:
    alias_lookup = {}

    for standard_field in STANDARD_FIELDS:
        alias_lookup[normalize_key(standard_field)] = standard_field

    for standard_field, aliases in COLUMN_ALIASES.items():
        alias_lookup[normalize_key(standard_field)] = standard_field

        for alias in aliases:
            alias_lookup[normalize_key(alias)] = standard_field
    
    return alias_lookup


def detect_column_mapping(headers: list[str]) -> dict[str, str]:
    alias_lookup = build_alias_lookup()

    result = {}

    normalized_headers = {
        normalize_key(header): header
        for header in headers
    }

    for normalized_header, original_header in normalized_headers.items():
        if normalized_header in alias_lookup:
            standard_field = alias_lookup[normalized_header]

            if standard_field not in result:
                result[standard_field] = original_header

    for normalized_header, original_header in normalized_headers.items():
        for normalized_alias, standard_field in alias_lookup.items():
            if standard_field in result:
                continue

            if not normalized_alias:
                continue

            if normalized_alias in normalized_header or normalized_header in normalized_alias:
                result[standard_field] = original_header

    return result


def get_missing_required_fields(column_mapping: dict[str, str]) -> list[str]:
    return [
        field 
        for field in REQUITRD_FIELDS
        if field not in column_mapping
    ]

def sniff_csv_dialect(path: Path):
    with open(path, mode="r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)

        try:
            return csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            return csv.excel
        

def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    dialect = sniff_csv_dialect(path)

    with open(path, mode="r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, dialect=dialect)

        headers = [
            clean_text(header)
            for header in (reader.fieldnames or [])
            if clean_text(header)
        ]

        rows = []

        for row in reader:
            rows.append(row)

    return headers, rows


def read_xlsx_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    sheet = workbook.active

    rows_iterator = sheet.iter_rows(values_only=True)

    try:
        first_row = next(rows_iterator)
    except StopIteration:
        return [], []
    
    headers = [
        clean_text(value)
        for value in first_row
        if clean_text(value)
    ]

    rows = []

    for values in rows_iterator:
        row = {}

        for index, header in enumerate(headers):
            value = values[index] if index < len(values) else ""
            row[header] = value
        
        rows.append(row)

    return headers, rows


def read_source_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return read_csv_rows(path)
    
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path)
    
    raise ValueError("Unsupported file format. Use CSV or XLSX")

def map_raw_row_to_standard_ticket(
        raw_row: dict[str, Any],
        column_mapping: dict[str, str],
        row_number: int,
)-> dict[str, str]:
    ticket = {}

    for standard_field in STANDARD_FIELDS:
        source_column = column_mapping.get(standard_field)

        if source_column:
            ticket[standard_field] = clean_text(raw_row.get(source_column, ""))
        else:
            ticket[standard_field] = ""

    if not ticket.get("ticket_id"):
        ticket["ticket_id"] = f"row_{row_number}"

    if not ticket.get("language"):
        ticket["language"] = "ru"

    if not ticket.get("category"):
        ticket["category"] = "general"

    return ticket

#map
def load_call_center_tickets(path: str | Path) -> tuple[list[dict[str, str]], dict[str, str], list[str]]:

    path = Path(path)

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    
    headers, raw_rows = read_source_rows(path)
    column_mapping = detect_column_mapping(headers)


    tickets = []

    for row_number, raw_row in enumerate(raw_rows, start=2):
        ticket = map_raw_row_to_standard_ticket(
            raw_row=raw_row,
            column_mapping=column_mapping,
            row_number=row_number,
        )
        tickets.append(ticket)

    return tickets, column_mapping, headers