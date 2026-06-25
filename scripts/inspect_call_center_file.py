#show col of files

#import libs

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

def inspect_csv(path: Path) -> list[str]:
    with open(path, mode="r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        return next(reader, [])
    

def inspect_xlsx(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active

    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])

    return [
        str(value).strip()
        for value in first_row
        if value is not None and str(value).strip()
    ]


def main() -> None:
    if len(sys.argv) <2:
        print("Usage: python scripts/inspect_call_center_file.py data/import/call_center_tickets.csv")
        return
    
    path = Path(sys.argv[1])

    if not path.exists():
        print(f"File not found: {path}")
        return
    
    suffix= path.suffix.lower()

    if suffix == ".csv":
        columns = inspect_csv(path)
    elif suffix in {".xlsx", ".xlsm"}:
        columns = inspect_xlsx(path)
    else:
        print("Unsupported file format. Use CSV or XLSX")
        return
    
    print('Detected columns:')
    for index, column in enumerate(columns, start=1):
        print(f"{index}. {column}")


if __name__ == "__main__":
    main()

